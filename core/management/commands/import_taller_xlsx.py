from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from flota.models import Colectivo, ParteDiario


def _parse_interno(value: str) -> Optional[int]:
    if not value:
        return None
    s = str(value)
    m = re.search(r"COCHE\s*(\d+)", s, re.IGNORECASE)
    if not m:
        m = re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _to_dt(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v))
    except Exception:
        return None


def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r"(\d+)", s.replace(".", "").replace(",", ""))
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


class Command(BaseCommand):
    help = "Importa REGISTRO DE TALLER (Google Forms) desde XLSX a flota.ParteDiario (tipo mantenimiento)."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Ruta al XLSX de REGISTRO DE TALLER (respuestas).")
        parser.add_argument("--dry-run", action="store_true", help="No escribe en DB; solo muestra resumen.")
        parser.add_argument("--limit", type=int, default=0, help="Limitar a N filas (0 = todas).")

    def handle(self, *args, **opts):
        p = Path(opts["path"])
        if not p.exists():
            raise CommandError(f"No existe el archivo: {p}")

        dry = bool(opts["dry_run"])
        limit = int(opts["limit"] or 0)

        wb = openpyxl.load_workbook(p.as_posix(), read_only=True, data_only=True)
        ws = wb.active

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if not headers or "Marca temporal" not in str(headers[0] or ""):
            raise CommandError("No reconozco el XLSX (encabezados). Asegurate de usar el export de Google Forms.")

        created = 0
        skipped = 0
        errors = 0

        def already_exists(colectivo_id: int, dt: datetime, desc: str) -> bool:
            return ParteDiario.objects.filter(colectivo_id=colectivo_id, fecha_evento=dt, descripcion=desc).exists()

        rows = ws.iter_rows(min_row=2, values_only=True)
        for idx, r in enumerate(rows, start=2):
            if limit and (idx - 1) > limit:
                break

            try:
                dt = _to_dt(r[0])
                coche = str(r[1] or "").strip()
                km = _to_int(r[2])
                chofer = str(r[3] or "").strip()
                sector = str(r[4] or "").strip()
                tipo_trab = str(r[5] or "").strip()
                desc = str(r[6] or "").strip()
                fotos_rep = str(r[7] or "").strip()
                registro = str(r[8] or "").strip()
                quien = str(r[9] or "").strip()

                interno = _parse_interno(coche)
                if not interno:
                    errors += 1
                    continue

                colectivo = Colectivo.objects.filter(interno=interno).first()
                if not colectivo:
                    errors += 1
                    continue

                if not desc:
                    desc = "Trabajo de taller (importado)"

                dt_local = dt
                if timezone.is_naive(dt_local):
                    dt_local = timezone.make_aware(dt_local, timezone.get_current_timezone())

                if already_exists(colectivo.id, dt_local, desc):
                    skipped += 1
                    continue

                obs_parts = []
                if sector:
                    obs_parts.append(f"Sector: {sector}")
                if tipo_trab:
                    obs_parts.append(f"Tipo trabajo: {tipo_trab}")
                if quien:
                    obs_parts.append(f"Realiza: {quien}")
                if fotos_rep:
                    obs_parts.append(f"Fotos repuesto: {fotos_rep}")
                if registro:
                    obs_parts.append(f"Registro visual: {registro}")

                observaciones = "IMPORT_XLSX: REGISTRO DE TALLER (Google Forms)"
                if obs_parts:
                    observaciones += "\n" + "\n".join(obs_parts)

                if not dry:
                    ParteDiario.objects.create(
                        colectivo=colectivo,
                        fecha_evento=dt_local,
                        reportado_por=None,
                        tipo=ParteDiario.Tipo.MANTENIMIENTO,
                        severidad=ParteDiario.Severidad.MEDIA,
                        estado=ParteDiario.Estado.RESUELTO,
                        odometro_km=km,
                        descripcion=desc,
                        observaciones=observaciones,
                        chofer_label=chofer,
                    )
                created += 1

            except Exception:
                errors += 1

        self.stdout.write(f"Archivo: {p.as_posix()}")
        self.stdout.write(f"Creados: {created} | Duplicados omitidos: {skipped} | Errores: {errors}")
        if dry:
            self.stdout.write("DRY-RUN: no se escribieron datos.")
        else:
            self.stdout.write(self.style.SUCCESS("OK: import terminado."))
