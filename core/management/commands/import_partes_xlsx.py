from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from flota.models import Colectivo, ParteDiario


User = get_user_model()


def _parse_interno(value: str) -> Optional[int]:
    if not value:
        return None
    s = str(value)
    m = re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _to_dt(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    try:
        # excel a veces trae string
        return datetime.fromisoformat(str(v))
    except Exception:
        return None


def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r"(\d+)", s.replace(".", "").replace(",", ""))
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _first_text(*vals) -> str:
    for v in vals:
        if v is None:
            continue
        t = str(v).strip()
        if t:
            return t
    return ""


@dataclass
class Row:
    dt: datetime
    email: str
    chofer: str
    km: Optional[int]
    coche: str
    mec: str
    elec: str
    car: str
    comb: str


class Command(BaseCommand):
    help = "Importa PARTES DIARIOS (Google Forms) desde XLSX a flota.ParteDiario (modo chofer)."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Ruta al XLSX de PARTES DIARIOS (respuestas).")
        parser.add_argument("--dry-run", action="store_true", help="No escribe en DB; solo muestra resumen.")
        parser.add_argument("--limit", type=int, default=0, help="Limitar a N filas (0 = todas).")

    def handle(self, *args, **opts):
        p = Path(opts["path"])
        if not p.exists():
            raise CommandError(f"No existe el archivo: {p}")

        dry = bool(opts["dry_run"])
        limit = int(opts["limit"] or 0)

        wb = openpyxl.load_workbook(p.as_posix(), read_only=True, data_only=True)
        ws = wb.active

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Se espera el orden exacto del export de Google Forms:
        # Marca temporal, Email, CHOFER, KILOMETRAJE, COCHE, PARTE MEC, PARTE ELEC, CARROCERIA, ADJUNTE FOTOS, COMBUSTIBLE RUTA
        if not headers or "Marca temporal" not in str(headers[0] or ""):
            raise CommandError("No reconozco el XLSX (encabezados). Asegurate de usar el export de Google Forms.")

        created = 0
        skipped = 0
        errors = 0

        def already_exists(colectivo_id: int, dt: datetime, desc: str) -> bool:
            return ParteDiario.objects.filter(colectivo_id=colectivo_id, fecha_evento=dt, descripcion=desc).exists()

        rows = ws.iter_rows(min_row=2, values_only=True)
        for idx, r in enumerate(rows, start=2):
            if limit and (idx - 1) > limit:
                break

            try:
                dt = _to_dt(r[0])
                email = str(r[1] or "").strip()
                chofer = str(r[2] or "").strip()
                km = _to_int(r[3])
                coche = str(r[4] or "").strip()

                mec = str(r[5] or "").strip()
                elec = str(r[6] or "").strip()
                car = str(r[7] or "").strip()
                comb = str(r[9] or "").strip()

                interno = _parse_interno(coche)
                if not interno:
                    errors += 1
                    continue

                colectivo = Colectivo.objects.filter(interno=interno).first()
                if not colectivo:
                    errors += 1
                    continue

                # descripcion obligatoria
                desc = _first_text(mec, elec, car, comb)
                if not desc:
                    desc = "Parte (importado)"

                # reportado_por por email si existe
                reporter = None
                if email:
                    reporter = User.objects.filter(email__iexact=email).first()

                dt_local = dt
                if timezone.is_naive(dt_local):
                    dt_local = timezone.make_aware(dt_local, timezone.get_current_timezone())

                if already_exists(colectivo.id, dt_local, desc):
                    skipped += 1
                    continue

                if not dry:
                    ParteDiario.objects.create(
                        colectivo=colectivo,
                        fecha_evento=dt_local,
                        reportado_por=reporter,
                        tipo=ParteDiario.Tipo.INCIDENCIA,
                        severidad=ParteDiario.Severidad.MEDIA,
                        estado=ParteDiario.Estado.ABIERTO,
                        odometro_km=km,
                        descripcion=desc,
                        observaciones="IMPORT_XLSX: PARTES DIARIOS (Google Forms)",
                        chofer_label=chofer,
                        parte_mecanico=mec,
                        parte_electrico=elec,
                        trabajos_carroceria_varios=car,
                        combustible_ruta_detalle=comb,
                    )
                created += 1

            except Exception:
                errors += 1

        self.stdout.write(f"Archivo: {p.as_posix()}")
        self.stdout.write(f"Creados: {created} | Duplicados omitidos: {skipped} | Errores: {errors}")
        if dry:
            self.stdout.write("DRY-RUN: no se escribieron datos.")
        else:
            self.stdout.write(self.style.SUCCESS("OK: import terminado."))
